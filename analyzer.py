import os
import json
import anthropic
import pandas as pd
import gradio as gr
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, get_column_letter

# ── 配置 ──────────────────────────────────────────────
REPORT_FOLDER = "./reports"
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
# ─────────────────────────────────────────────────────

client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

def read_excel(filepath):
    """读取 Excel 文件，返回所有 sheet 的数据摘要"""
    xl = pd.ExcelFile(filepath)
    sheets = {}
    for sheet in xl.sheet_names:
        df = xl.parse(sheet)
        df = df.dropna(how="all").dropna(axis=1, how="all")
        if df.empty:
            continue
        sheets[sheet] = {
            "columns": list(df.columns),
            "rows": len(df),
            "preview": df.head(20).to_string(index=False),
            "stats": df.describe(include="all").to_string() if not df.empty else ""
        }
    return sheets

def analyze_with_claude(filename, sheets_data):
    """调用 Claude API 分析数据"""
    content_parts = [f"文件名：{filename}\n"]
    for sheet_name, data in sheets_data.items():
        content_parts.append(f"\n=== Sheet：{sheet_name} ===")
        content_parts.append(f"字段：{', '.join(str(c) for c in data['columns'])}")
        content_parts.append(f"数据行数：{data['rows']}")
        content_parts.append(f"\n数据预览：\n{data['preview']}")
        if data['stats']:
            content_parts.append(f"\n统计摘要：\n{data['stats']}")
    
    user_content = "\n".join(content_parts)

    message = client.messages.create(
        model="claude-3-5-sonnet-20240620", 
        max_tokens=2000,
        system="""你是一位专业的水产养殖实验数据分析师，专门分析小龙虾养殖实验数据。
请用中文分析实验数据，输出结构如下（直接输出文本，不用 Markdown）：

【数据概述】
简要描述数据内容、实验类型、数据规模。

【关键指标分析】
逐一分析重要指标的数值范围、均值、趋势。

【异常与风险】
指出数据中的异常值、缺失值或值得关注的波动。

【实验结论】
基于数据得出主要结论，哪组/哪个条件表现最好。

【养殖建议】
根据分析结果，给出具体可操作的养殖改进建议。

语言简洁专业，每个部分 3-5 条要点。""",
        messages=[
            {"role": "user", "content": f"请分析以下龙虾实验数据：\n\n{user_content}"}
        ]
    )
    return message.content[0].text

def save_report(filename, analysis_text, sheets_data, original_filepath):
    """将分析结果保存为 Excel 报告"""
    Path(REPORT_FOLDER).mkdir(parents=True, exist_ok=True)
    report_name = f"报告_{Path(filename).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    report_path = Path(REPORT_FOLDER) / report_name

    wb = __import__('openpyxl').Workbook()
    ws = wb.active
    ws.title = "分析报告"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80

    header_fill = PatternFill("solid", fgColor="1D6E4F")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=14)
    body_font   = Font(name="微软雅黑", size=10)
    wrap_align  = Alignment(wrap_text=True, vertical="top")

    ws.merge_cells("A1:B1")
    ws["A1"] = f"龙虾实验数据分析报告 — {Path(filename).stem}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["A2"] = "生成时间"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    sections = analysis_text.split("【")
    for sec in sections:
        if "】" in sec:
            title, content = sec.split("】", 1)
            row = ws.max_row + 2
            ws.cell(row=row, column=1, value=f"【{title}】").font = Font(name="微软雅黑", bold=True, size=11, color="1D6E4F")
            ws.cell(row=row, column=2, value=content.strip()).font = body_font
            ws.cell(row=row, column=2).alignment = wrap_align

    # 复制原始数据
    for sheet_name, data in sheets_data.items():
        try:
            df = pd.read_excel(original_filepath, sheet_name=sheet_name)
            ws2 = wb.create_sheet(title=f"数据_{sheet_name}"[:31])
            for col_idx, col_name in enumerate(df.columns, 1):
                ws2.cell(row=1, column=col_idx, value=str(col_name)).fill = PatternFill("solid", fgColor="2E7D5A")
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, val in enumerate(row, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=val)
        except Exception:
            pass

    wb.save(report_path)
    return report_path

def process_file_ui(file_obj):
    if file_obj is None:
        return "请先上传 Excel 文件", None
    
    if not ANTHROPIC_API_KEY:
        return "❌ 错误：缺少 Claude API 密钥，请在环境变量中配置 ANTHROPIC_API_KEY", None

    filepath = file_obj.name
    filename = Path(filepath).name
    
    try:
        sheets_data = read_excel(filepath)
        if not sheets_data:
            return "文件为空或读取失败", None
            
        analysis = analyze_with_claude(filename, sheets_data)
        report_path = save_report(filename, analysis, sheets_data, filepath)
        return "✅ 处理成功！请在右侧下载报告。", str(report_path)
    except Exception as e:
        return f"❌ 处理失败: {str(e)}", None

# 构建网页界面
demo = gr.Interface(
    fn=process_file_ui,
    inputs=gr.File(label="📥 上传龙虾实验数据 (.xlsx)"),
    outputs=[
        gr.Textbox(label="运行状态"), 
        gr.File(label="📤 下载 AI 分析报告")
    ],
    title="🦞 24小时 AI 龙虾数据分析助手",
    description="上传你的 Excel 数据，Claude 会自动帮你分析并生成报告。"
)

if __name__ == "__main__":
    # Railway 会自动分配端口给 PORT 环境变量
    port = int(os.getenv("PORT", 7860))
    demo.launch(server_name="0.0.0.0", server_port=port)
